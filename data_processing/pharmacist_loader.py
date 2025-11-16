# pharmacy_deserts/data_processing/pharmacist_loader.py
import pandas as pd
import openpyxl
from pathlib import Path

# Use relative import to avoid path issues
try:
    from utils.cache import cache_data
except ImportError:
    # If relative import fails, try absolute or provide fallback
    try:
        import sys
        from pathlib import Path
        sys.path.insert(0, str(Path(__file__).parent.parent))
        from utils.cache import cache_data
    except:
        # Fallback: no-op decorator
        def cache_data(func):
            return func

@cache_data
def load_all_pharmacist_data(data_dir="data"):
    """
    Load and combine all subset pharmacist files (.xlsm format).
    Returns DataFrame with Short_ZIP, Combined, Award, Phone, and Address columns.
    """
    # Look for .xlsm files instead of .xlsx
    subset_files = sorted(Path(data_dir).glob("subset*_Table2_filter.xlsm"))
    
    if not subset_files:
        print("Warning: No subset pharmacist .xlsm files found")
        return pd.DataFrame(columns=['Short_ZIP', 'Combined', 'Award', 'Phone', 'Address'])
    
    print(f"Loading {len(subset_files)} pharmacist subset files...")
    
    all_data = []
    for file in subset_files:
        try:
            # Load the workbook
            xlsx = openpyxl.load_workbook(file, read_only=True, data_only=True)
            
            if not xlsx.sheetnames:
                print(f"  ✗ Skipped {file.name}: No sheets found")
                continue
            
            # Get first sheet data
            sheet = xlsx[xlsx.sheetnames[0]]
            data = list(sheet.values)
            
            if not data:
                print(f"  ✗ Skipped {file.name}: Empty sheet")
                continue
            
            # Create DataFrame with first row as headers
            df = pd.DataFrame(data[1:], columns=data[0])
            
            # Check for required columns (handle both "Short_ ZIP" and "Shorter ZIP")
            # Note: The actual column names have spaces in them!
            zip_col = None
            if 'Short_ ZIP' in df.columns:
                zip_col = 'Short_ ZIP'
            elif 'Shorter ZIP' in df.columns:
                zip_col = 'Shorter ZIP'
            elif 'Short_ZIP' in df.columns:
                zip_col = 'Short_ZIP'
            
            # Check for Award column
            award_col = 'Award?' if 'Award?' in df.columns else None
            
            # Check for Phone and Address columns
            phone_col = 'Provider Business Practice Location Address Telephone Number' if 'Provider Business Practice Location Address Telephone Number' in df.columns else None
            address_col = 'Provider First Line Business Practice Location Address' if 'Provider First Line Business Practice Location Address' in df.columns else None
            
            if zip_col and 'Combined' in df.columns:
                # Extract columns (include Award, Phone, Address if available)
                cols_to_extract = [zip_col, 'Combined']
                new_cols = ['Short_ZIP', 'Combined']
                
                if award_col:
                    cols_to_extract.append(award_col)
                    new_cols.append('Award')
                if phone_col:
                    cols_to_extract.append(phone_col)
                    new_cols.append('Phone')
                if address_col:
                    cols_to_extract.append(address_col)
                    new_cols.append('Address')
                
                df_subset = df[cols_to_extract].copy()
                df_subset.columns = new_cols
                
                # Add missing columns with None
                if 'Award' not in df_subset.columns:
                    df_subset['Award'] = None
                if 'Phone' not in df_subset.columns:
                    df_subset['Phone'] = None
                if 'Address' not in df_subset.columns:
                    df_subset['Address'] = None
                
                all_data.append(df_subset)
                
                extras = []
                if award_col:
                    extras.append("Award")
                if phone_col:
                    extras.append("Phone")
                if address_col:
                    extras.append("Address")
                extra_info = f" (with {', '.join(extras)})" if extras else ""
                print(f"  ✓ Loaded {file.name}: {len(df)} records{extra_info}")
            else:
                print(f"  ✗ Skipped {file.name}: Missing required columns")
                print(f"    Available: {df.columns.tolist()[:10]}")
                
        except Exception as e:
            print(f"  ✗ Error loading {file.name}: {e}")
            continue
    
    if not all_data:
        print("Warning: No valid pharmacist data found")
        return pd.DataFrame(columns=['Short_ZIP', 'Combined', 'Award', 'Phone', 'Address'])
    
    # Combine all subsets
    combined = pd.concat(all_data, ignore_index=True)
    
    # Remove any null values in required columns
    combined = combined.dropna(subset=['Short_ZIP', 'Combined'])
    
    # Count awarded pharmacists (Award=1 means they received an award)
    awarded_count = (combined['Award'] == 1).sum() if 'Award' in combined.columns else 0
    
    print(f"Total pharmacist records loaded: {len(combined)}")
    print(f"Unique ZIPs with pharmacists: {combined['Short_ZIP'].nunique()}")
    print(f"Pharmacists with awards: {awarded_count}")
    
    return combined

@cache_data
def get_pharmacists_for_zip(zip_code, pharmacist_df):
    """
    Get all pharmacists in a specific ZIP code.
    Returns list of tuples: (pharmacist_name, has_award, phone, address)
    """
    if pharmacist_df is None or pharmacist_df.empty:
        return []
    
    # Convert to string for matching
    zip_str = str(zip_code).strip()
    
    # Filter for this ZIP
    zip_data = pharmacist_df[pharmacist_df['Short_ZIP'].astype(str).str.strip() == zip_str]
    
    if zip_data.empty:
        return []
    
    # Build list of (name, has_award, phone, address) tuples
    pharmacists = []
    for _, row in zip_data.iterrows():
        name = row['Combined']
        if pd.notna(name) and str(name).strip():
            # Check if they have an award (Award column exists and value is 1)
            has_award = False
            if 'Award' in row.index and pd.notna(row['Award']):
                # Award=1 means they received an award, Award=0 means no award
                has_award = (int(row['Award']) == 1)
            
            # Get phone and address
            phone = str(row.get('Phone', '')).strip() if pd.notna(row.get('Phone')) else ''
            address = str(row.get('Address', '')).strip() if pd.notna(row.get('Address')) else ''
            
            # Format phone number nicely if it exists
            if phone and phone not in ['nan', 'None', '']:
                # Remove any non-digit characters (handles both string and float formats)
                phone_digits = ''.join(filter(str.isdigit, phone))
                if len(phone_digits) == 10:
                    phone = f"({phone_digits[:3]}) {phone_digits[3:6]}-{phone_digits[6:]}"
                elif len(phone_digits) > 10:
                    # Handle cases with country code or extra digits
                    phone = f"({phone_digits[-10:-7]}) {phone_digits[-7:-4]}-{phone_digits[-4:]}"
                else:
                    # Keep as is if not 10 digits
                    phone = phone_digits if phone_digits else ''
            else:
                phone = ''
            
            pharmacists.append((str(name).strip(), has_award, phone, address))
    
    # Sort: award winners first, then alphabetically by name
    pharmacists.sort(key=lambda x: (not x[1], x[0].lower()))
    
    return pharmacists
